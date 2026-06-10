#!/usr/bin/env python3
"""
LIVE Market Dashboard (Streamlit) - 9-section management layout
===============================================================
Auto-refreshes. Every number comes from a real source and is labelled.
Items with a free feed update automatically; items without a free source
show a blank "-" and are never invented. Add a free FRED key in the sidebar
to unlock the US/EU/India macro rows.

RUN:  python3 -m streamlit run app.py
"""

import datetime as dt
import html as _html
import io
import json
import os
import re

import pandas as pd
import requests
import streamlit as st

try:
    from streamlit_autorefresh import st_autorefresh
    HAS_AUTOREFRESH = True
except Exception:
    HAS_AUTOREFRESH = False

try:
    import yfinance as yf
    HAS_YF = True
except Exception:
    HAS_YF = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    HAS_XLSX = True
except Exception:
    HAS_XLSX = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, Image)
    HAS_PDF = True
except Exception:
    HAS_PDF = False

H = {"User-Agent": "Mozilla/5.0 (personal market dashboard)"}
MANUAL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "manual_data.json")
# Local timezone label, correct on any machine/OS (e.g. IST, EST, GMT)
TZ = dt.datetime.now().astimezone().tzname() or "local"



def load_manual():
    try:
        with open(MANUAL_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def save_manual(d):
    """Merge the given keys into manual_data.json (keeps other saved fields)."""
    try:
        cur = load_manual()
        cur.update(d)
        with open(MANUAL_FILE, "w", encoding="utf-8") as f:
            json.dump(cur, f, indent=2)
        return True
    except Exception:
        return False


manual = load_manual()
FRED_API_KEY = manual.get("fred_key", "").strip()

# Latest published India macro that has NO free real-time API (released only in
# MoSPI/PIB press releases). These are the defaults shown until you override them
# in the sidebar ("India macro" box). Update when a new release comes out:
#   GDP  -> quarterly by MoSPI (NSO), ~last business day of the 2nd month after
#   WPI  -> monthly by the Office of the Economic Adviser, ~14th of next month
LATEST_INDIA_GDP = ("7.70", "FY26")        # full-year FY26 real GDP growth %
LATEST_INDIA_WPI = ("8.30", "Apr 2026")    # WPI inflation, YoY %

# ---------------------------------------------------------------------------
# TICKERS (Yahoo Finance)
# ---------------------------------------------------------------------------
EQUITIES = [
    ("India", "Nifty 50", "^NSEI"), ("India", "Sensex", "^BSESN"),
    ("US", "S&P 500", "^GSPC"), ("US", "Nasdaq", "^IXIC"),
    ("US", "Dow Jones", "^DJI"), ("Germany", "DAX", "^GDAXI"),
    ("UK", "FTSE 100", "^FTSE"), ("China", "Shanghai Composite", "000001.SS"),
    ("Hong Kong", "Hang Seng", "^HSI"), ("Japan", "Nikkei 225", "^N225"),
]
FX = [("USD/INR", "INR=X"), ("EUR/INR", "EURINR=X"), ("GBP/INR", "GBPINR=X"),
      ("JPY/INR", "JPYINR=X"), ("USD/CNY", "CNY=X")]
_OZ_G = 31.1035       # grams per troy ounce
_LB_KG = 0.453592     # kg per pound
# (label, yahoo ticker, factor to convert the USD native-unit price to the unit)
COMMODITIES = [
    ("Gold 24K (INR/10g)",      "GC=F", 10 / _OZ_G),
    ("Silver (INR/kg)",         "SI=F", 1000 / _OZ_G),
    ("Brent Crude (INR/bbl)",   "BZ=F", 1.0),
    ("WTI Crude (INR/bbl)",     "CL=F", 1.0),
    ("Natural Gas (INR/MMBtu)", "NG=F", 1.0),
    ("Copper (INR/kg)",         "HG=F", 1 / _LB_KG),
    ("Copper (INR/tonne)",      "HG=F", 1000 / _LB_KG),
    ("Aluminum (INR/t)",        "ALI=F", 1.0),
]
VOL = [("India VIX", "^INDIAVIX"), ("CBOE VIX", "^VIX")]
REITS = [("Embassy REIT", "EMBASSY.BO"), ("Mindspace REIT", "MINDSPACE.BO"),
         ("Brookfield REIT", "BIRET.BO"), ("Nexus Select REIT", "NXST.BO")]
INVITS = [("IndiGrid InvIT", "INDIGRID.BO"), ("Powergrid InvIT", "PGINVIT.BO"),
          ("IRB InvIT", "IRBINVIT.BO")]
CRYPTO = [("Bitcoin", "BTC-USD"), ("Ethereum", "ETH-USD")]



# ---------------------------------------------------------------------------
# FETCHERS
# ---------------------------------------------------------------------------
@st.cache_data(ttl=300)
def yahoo(tickers):
    out = {}
    if not HAS_YF:
        return out
    for item in tickers:
        tk = item[-1]
        try:
            h = yf.Ticker(tk).history(period="6y")["Close"].dropna()
            if len(h) < 2:
                h = yf.Ticker(tk).history(period="5d")["Close"].dropna()
            if len(h) >= 2:
                p, prev = float(h.iloc[-1]), float(h.iloc[-2])
                rec = {"price": p, "d1": (p/prev-1)*100, "base": {"d1": prev}}
                # YTD: first close of the current calendar year
                this_year = h[h.index.year == h.index[-1].year]
                ys = float(this_year.iloc[0]) if len(this_year) else float(h.iloc[0])
                rec["ytd"] = (p/ys-1)*100
                rec["base"]["ytd"] = ys
                # 1M / 3M / 6M / 1Y / 5Y: close nearest to N days ago
                last_date = h.index[-1]
                for label, days in (("m1", 30), ("m3", 91), ("m6", 182),
                                    ("y1", 365), ("y5", 1826)):
                    cutoff = last_date - pd.Timedelta(days=days)
                    past = h[h.index <= cutoff]
                    base = float(past.iloc[-1]) if len(past) else None
                    rec[label] = (p/base-1)*100 if base else None
                    rec["base"][label] = base
                out[tk] = rec
        except Exception:
            pass
    return out


@st.cache_data(ttl=21600)  # 6h — distributions are announced quarterly
def ttm_payout(tk):
    """Sum of the ACTUAL distributions Yahoo recorded in the trailing 12 months.
    Real data imported from Yahoo, not an assumption. Returns rupees/unit, or
    None if Yahoo has no distribution history for this ticker."""
    if not HAS_YF:
        return None
    try:
        s = yf.Ticker(tk).dividends            # pandas Series, indexed by pay date
        if s is None or len(s) == 0:
            return None
        now = pd.Timestamp.now(tz=s.index.tz) if s.index.tz is not None \
            else pd.Timestamp.now()
        total = float(s[s.index >= now - pd.Timedelta(days=365)].sum())
        return total if total > 0 else None
    except Exception:
        return None


@st.cache_data(ttl=300)
def treasury():
    yr = dt.date.today().year
    url = (f"https://home.treasury.gov/resource-center/data-chart-center/"
           f"interest-rates/daily-treasury-rates.csv/{yr}/all"
           f"?type=daily_treasury_yield_curve&field_tdr_date_value={yr}"
           f"&page&_format=csv")
    r = requests.get(url, headers=H, timeout=20); r.raise_for_status()
    df = pd.read_csv(io.StringIO(r.text))
    df["Date"] = pd.to_datetime(df["Date"]); df = df.sort_values("Date")
    a, b = df.iloc[-1], df.iloc[-2]
    g = lambda c: (float(a[c]), float(a[c]-b[c])) if c in df.columns else (None, None)
    return {"1Y": g("1 Yr"), "2Y": g("2 Yr"), "10Y": g("10 Yr"),
            "date": a["Date"].date().isoformat()}


@st.cache_data(ttl=3600)  # 1h — JGB curve updates once daily
def japan_jgb():
    """Japan JGB yields from the Ministry of Finance daily CSV (free, plain HTTP,
    works on Streamlit Cloud — no browser needed). The CSV has a full curve:
    Date,1Y,2Y,3Y,...,10Y,...,40Y. Returns
    {'1Y':float,'2Y':float,'10Y':float,'10Y_chg':float(pct-points)} or {}."""
    url = ("https://www.mof.go.jp/english/policy/jgbs/reference/"
           "interest_rate/jgbcme.csv")
    try:
        r = requests.get(url, headers=H, timeout=20)
        if r.status_code != 200 or not r.text.strip():
            return {}
        lines = [ln for ln in r.text.splitlines() if ln.strip()]
        hdr_i = next((i for i, l in enumerate(lines)
                      if l.split(",")[0].strip() == "Date"), None)
        if hdr_i is None:
            return {}
        header = [h.strip() for h in lines[hdr_i].split(",")]
        idx = {name: j for j, name in enumerate(header)}
        # Keep only real data rows (first cell is a date like 2026/5/18). The CSV
        # ends with footer junk (an all-commas line and a "clear your cache"
        # message) — those must be skipped or the latest row parses as empty.
        rows = [l.split(",") for l in lines[hdr_i + 1:]
                if re.match(r"^\s*\d{4}/\d{1,2}/\d{1,2}", l.split(",")[0])]
        if not rows:
            return {}

        def val(row, col):
            try:
                return float(row[idx[col]])
            except Exception:
                return None

        last = rows[-1]
        out = {"1Y": val(last, "1Y"), "2Y": val(last, "2Y"),
               "10Y": val(last, "10Y")}
        if len(rows) >= 2 and out["10Y"] is not None:
            p10 = val(rows[-2], "10Y")
            if p10 is not None:
                out["10Y_chg"] = round(out["10Y"] - p10, 3)
        return out
    except Exception:
        return {}


@st.cache_data(ttl=3600)  # 1h — euro-area curve updates daily ~noon CET
def ecb_yield():
    """Euro-area AAA government bond yields (the German-bund benchmark curve)
    from the ECB Data Portal SDMX REST API — free, plain HTTP, works on cloud.
    Returns {'1Y':float,'2Y':float,'10Y':float} or {}. Used for 'Germany'."""
    base = ("https://data-api.ecb.europa.eu/service/data/YC/"
            "B.U2.EUR.4F.G_N_A.SV_C_YM.SR_{}?lastNObservations=1&format=csvdata")
    out = {}
    for tenor in ("1Y", "2Y", "10Y"):
        try:
            r = requests.get(base.format(tenor), headers=H, timeout=20)
            if r.status_code != 200 or not r.text.strip():
                continue
            lines = [ln for ln in r.text.splitlines() if ln.strip()]
            if len(lines) < 2:
                continue
            cols = lines[0].split(",")
            vi = cols.index("OBS_VALUE") if "OBS_VALUE" in cols else -1
            last = lines[-1].split(",")
            if vi >= 0 and vi < len(last):
                out[tenor] = round(float(last[vi]), 2)
        except Exception:
            pass
    return out


@st.cache_data(ttl=120)
def crypto():
    url = ("https://api.coingecko.com/api/v3/simple/price?ids=bitcoin,ethereum"
           "&vs_currencies=usd&include_24hr_change=true")
    r = requests.get(url, headers=H, timeout=20); r.raise_for_status()
    d = r.json()
    return {"Bitcoin": (d["bitcoin"]["usd"], d["bitcoin"]["usd_24h_change"]),
            "Ethereum": (d["ethereum"]["usd"], d["ethereum"]["usd_24h_change"])}


@st.cache_data(ttl=1800)  # 30 min — Indian gold rate updates a few times a day
def india_gold():
    """LIVE Indian retail 24K gold rate in INR per gram, scraped from Goodreturns
    (server-rendered, no JS). This is the actual Indian quoted rate — it INCLUDES
    the duty/premium that the COMEX-spot conversion omits, so it matches what
    jewellers and other Indian sites show. Returns {'24k': float} per gram, or {}."""
    srcs = ["https://www.goodreturns.in/gold-rates/",
            "https://www.goodreturns.in/gold-rates/mumbai.html",
            "https://www.goodreturns.in/gold-rates/delhi.html"]
    for url in srcs:
        try:
            r = requests.get(url, headers=_BROWSER_H, timeout=15)
            if r.status_code != 200:
                continue
            txt = re.sub(r"\s+", " ",
                         _html.unescape(_TAGS.sub(" ", _STRIP.sub(" ", r.text))))
            m24 = re.search(r"([\d,]{3,})\s*per\s*gram\s*for\s*24", txt, re.I)
            g24 = float(m24.group(1).replace(",", "")) if m24 else None
            # sanity: Indian 24K/gram is in the thousands (₹), reject junk
            if g24 and 1000 <= g24 <= 100000:
                return {"24k": g24}
        except Exception:
            pass
    return {}


@st.cache_data(ttl=600)
def fred(series, key):
    if not key:
        return None
    try:
        url = (f"https://api.stlouisfed.org/fred/series/observations?series_id="
               f"{series}&api_key={key}&file_type=json&sort_order=desc&limit=1")
        v = requests.get(url, timeout=20).json()["observations"][0]["value"]
        return float(v)          # FRED sends "." for missing -> ValueError -> None
    except Exception:
        return None


@st.cache_data(ttl=21600)  # 6h — monthly macro series
def fred_latest(series, key):
    """Latest (value, 'Mon YYYY') for a FRED series, skipping missing '.' obs.
    Used for current monthly macro figures. Returns (float, str) or (None, None)."""
    if not key:
        return None, None
    try:
        url = (f"https://api.stlouisfed.org/fred/series/observations?series_id="
               f"{series}&api_key={key}&file_type=json&sort_order=desc&limit=12")
        obs = requests.get(url, timeout=20).json().get("observations", [])
        for o in obs:
            v = o.get("value")
            if v not in (".", "", None):
                d = o.get("date", "")
                try:
                    label = dt.datetime.strptime(d, "%Y-%m-%d").strftime("%b %Y")
                except Exception:
                    label = d
                return float(v), label
        return None, None
    except Exception:
        return None, None


@st.cache_data(ttl=21600)  # 6h — historical curve barely changes intraday
def fred_history(series, key, years=5):
    """Full observation history for a FRED series over the last `years` years.
    Returns a date-indexed pandas Series of floats (missing '.' values dropped),
    or None. Daily series (e.g. DGS10) come back daily; OECD cross-country
    series come back monthly. Used by the 5-year 10Y yield chart."""
    if not key:
        return None
    try:
        start = (dt.date.today() - dt.timedelta(days=365 * years + 10)).isoformat()
        url = (f"https://api.stlouisfed.org/fred/series/observations?series_id="
               f"{series}&api_key={key}&file_type=json&sort_order=asc"
               f"&observation_start={start}")
        obs = requests.get(url, timeout=25).json().get("observations", [])
        pairs = [(o["date"], float(o["value"])) for o in obs
                 if o.get("value") not in (".", "", None)]
        if not pairs:
            return None
        s = pd.Series({pd.to_datetime(d): v for d, v in pairs}).sort_index()
        return s if len(s) else None
    except Exception:
        return None


@st.cache_data(ttl=86400)
def bis_policy(cc):
    """Central bank policy rate from BIS (free, weekly). cc = IN, GB, XM, US.
    Best-effort across a few endpoint shapes; returns '5.25%' style or None."""
    urls = [
        f"https://stats.bis.org/api/v1/data/WS_CBPOL/D.{cc}/all?lastNObservations=1&format=csv",
        f"https://stats.bis.org/api/v1/data/WS_CBPOL/M.{cc}/all?lastNObservations=1&format=csv",
        f"https://stats.bis.org/api/v1/data/BIS,WS_CBPOL,1.0/D.{cc}?lastNObservations=1&format=csv",
        f"https://stats.bis.org/api/v1/data/BIS,WS_CBPOL,1.0/M.{cc}?lastNObservations=1&format=csv",
    ]
    for url in urls:
        try:
            r = requests.get(url, headers=H, timeout=20)
            if r.status_code == 200 and r.text.strip():
                df = pd.read_csv(io.StringIO(r.text))
                vcol = [c for c in df.columns if c.upper() == "OBS_VALUE"]
                if vcol:
                    val = pd.to_numeric(df[vcol[0]], errors="coerce").dropna()
                    if len(val):
                        return f"{float(val.iloc[-1]):.2f}%"
        except Exception:
            pass
    return None


@st.cache_data(ttl=3600)
def worldbank(indicator):
    """India macro from the World Bank API (free, no key). Returns (value, year).
    Scans the last 10 years newest-first and retries once, so a transient miss
    doesn't get stuck."""
    for _ in range(2):
        try:
            url = (f"https://api.worldbank.org/v2/country/IND/indicator/{indicator}"
                   f"?format=json&mrv=10")
            r = requests.get(url, headers=H, timeout=25); r.raise_for_status()
            data = r.json()
            if isinstance(data, list) and len(data) > 1 and data[1]:
                for obs in data[1]:  # newest first
                    if obs.get("value") is not None:
                        return float(obs["value"]), obs.get("date")
        except Exception:
            pass
    return None


# ---------------------------------------------------------------------------
# BANK FD SCRAPER (best-effort; only banks that serve rates in raw HTML)
# ---------------------------------------------------------------------------
_BROWSER_H = {
    "User-Agent": ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/124.0 Safari/537.36"),
    "Accept-Language": "en-US,en;q=0.9",
}
_STRIP = re.compile(r"<script.*?</script>|<style.*?</style>", re.S | re.I)
_TAGS = re.compile(r"<[^>]+>")
# Every "<n> <unit> ... <rate>%" pair on the page.
_TENOR_RATE = re.compile(
    r"(\d{1,4})\s*(day|days|month|months|year|years|yr|yrs)\b[^%]{0,45}?"
    r"([3-9]\.\d{1,2})\s*%", re.I)
_UNIT_DAYS = {"day": 1, "days": 1, "month": 30, "months": 30,
              "year": 365, "years": 365, "yr": 365, "yrs": 365}


@st.cache_data(ttl=21600)  # 6h — FD rates change only a few times a year
def bank_fd(url):
    """Scrape a bank's ~1-year FD rate from its public page via a simple fetch.
    Works for banks that serve rates in raw HTML (HDFC, Axis)."""
    try:
        r = requests.get(url, headers=_BROWSER_H, timeout=12)
        if r.status_code != 200:
            return None
        text = _html.unescape(_TAGS.sub(" ", _STRIP.sub(" ", r.text)))
        return _pick_1y(re.sub(r"\s+", " ", text))
    except Exception:
        return None


def _pick_1y(text):
    """From visible text, collect every (tenor->days, rate) pair and return the
    rate for the bucket nearest 365 days. Shared by the simple fetch and the
    headless browser. Returns {rate, days, cands} or None."""
    cands = []
    for m in _TENOR_RATE.finditer(text):
        n, unit, rate = int(m.group(1)), m.group(2).lower(), float(m.group(3))
        days = n * _UNIT_DAYS[unit]
        if 2.0 <= rate <= 9.5 and 7 <= days <= 3700:
            cands.append((days, rate))
    if not cands:
        return None
    best = min(cands, key=lambda c: (abs(c[0] - 365), c[0]))
    seen, sample = set(), []
    for d, rt in cands:
        if (d, rt) not in seen:
            seen.add((d, rt)); sample.append((d, rt))
        if len(sample) >= 12:
            break
    return {"rate": best[1], "days": best[0], "cands": sample}


@st.cache_data(ttl=21600)  # 6h — heavy call, so cache hard
def bank_fd_browser(url):
    """Scrape a JS-rendered bank page (SBI, ICICI, Kotak) using a real Chrome
    via Playwright: open the page, let the rate JS run, then parse the text.
    Returns {rate, days, cands} or None. Needs: pip install playwright;
    python -m playwright install chromium."""
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return None
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(user_agent=_BROWSER_H["User-Agent"])
            page.goto(url, timeout=45000, wait_until="domcontentloaded")
            page.wait_for_timeout(6000)        # let the rate table load
            body = page.inner_text("body")
            browser.close()
        return _pick_1y(re.sub(r"\s+", " ", body))
    except Exception:
        return None


# ---------------------------------------------------------------------------
# MULTI-TENOR FD SCRAPER
# ---------------------------------------------------------------------------
_FD_TARGETS = {"1M": 30, "3M": 91, "6M": 182, "1Y": 365, "3Y": 1095}
_FD_TOL     = {"1M": 25, "3M": 50, "6M": 75,  "1Y": 120, "3Y": 250}

# Wider regex: rate range 2.x–9.x, longer gap so wide tables still match.
_TENOR_RATE_WIDE = re.compile(
    r"(\d{1,4})\s*(day|days|month|months|year|years|yr|yrs|mo|mth|mths)\b"
    r"[^%]{0,80}?([2-9]\.\d{1,2})\s*%", re.I)
_UNIT_DAYS_W = {"day": 1, "days": 1, "month": 30, "months": 30,
                "mo": 30, "mth": 30, "mths": 30,
                "year": 365, "years": 365, "yr": 365, "yrs": 365}


def _pick_all_tenors(text):
    """Collect every (tenor->days, rate) pair from page text and bucket them
    into the 5 target tenors. The FIRST rate after a tenor is taken (aggregator
    tables list general-public before senior-citizen), so we read the general
    public rate. Returns {tenor_label: 'X.XX'}."""
    cands = []
    for m in _TENOR_RATE_WIDE.finditer(text):
        n = int(m.group(1))
        unit = m.group(2).lower()
        rate = float(m.group(3))
        days = n * _UNIT_DAYS_W.get(unit, 0)
        if 2.0 <= rate <= 9.75 and 7 <= days <= 4000:
            cands.append((days, rate))
    result = {}
    for label, want in _FD_TARGETS.items():
        tol = _FD_TOL[label]
        near = [c for c in cands if abs(c[0] - want) <= tol]
        if near:
            best = min(near, key=lambda c: abs(c[0] - want))
            result[label] = f"{best[1]:.2f}"
    return result


def _merge_tenors(dst, src):
    """Fill only the tenors dst is still missing — earlier (higher-priority)
    sources win, later sources just patch the gaps."""
    for k, v in src.items():
        dst.setdefault(k, v)
    return dst


def _scrape_http(url):
    """Plain HTTP fetch + parse. Thread-safe. Returns {tenor: rate} or {}."""
    if not url:
        return {}
    try:
        r = requests.get(url, headers=_BROWSER_H, timeout=15)
        if r.status_code == 200:
            txt = _html.unescape(_TAGS.sub(" ", _STRIP.sub(" ", r.text)))
            return _pick_all_tenors(re.sub(r"\s+", " ", txt))
    except Exception:
        pass
    return {}


def _scrape_browser(url):
    """Headless-browser fetch + parse. MAIN THREAD ONLY (Playwright sync API
    cannot run inside a worker thread). Returns {tenor: rate} or {}."""
    if not url:
        return {}
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(user_agent=_BROWSER_H["User-Agent"])
            page.goto(url, timeout=45000, wait_until="domcontentloaded")
            page.wait_for_timeout(7000)          # let rate JS finish
            try:                                  # nudge lazy tables into view
                page.mouse.wheel(0, 4000)
                page.wait_for_timeout(1500)
            except Exception:
                pass
            body = page.inner_text("body")
            browser.close()
        return _pick_all_tenors(re.sub(r"\s+", " ", body))
    except Exception:
        return {}


# Per bank: (name, [aggregator URLs in priority order], direct bank URL).
# BankBazaar pages are server-rendered with full tenure tables (most reliable),
# so they go first; PolicyBazaar (verified "-fd-rates" slug) and Paisabazaar
# follow. Results are MERGED across all of them, so a tenor missing on one site
# is filled by another. 404s on a guessed URL are harmless (ignored).
_FD_INSTITUTIONS = [
    ("SBI", [
        "https://www.bankbazaar.com/fixed-deposit/sbi-fixed-deposit-rate.html",
        "https://www.policybazaar.com/fd-interest-rates/sbi-fd-rates/",
        "https://www.paisabazaar.com/fixed-deposit/sbi-fd-rates/",
    ], "https://sbi.co.in/web/interest-rates/deposit-rates/retail-domestic-term-deposits"),
    ("HDFC Bank", [
        "https://www.bankbazaar.com/fixed-deposit/hdfc-fixed-deposit-rate.html",
        "https://www.policybazaar.com/fd-interest-rates/hdfc-bank-fd-rates/",
        "https://www.paisabazaar.com/fixed-deposit/hdfc-bank-fd-rates/",
    ], "https://www.hdfcbank.com/personal/save-invest/deposits/fixed-deposit/fixed-deposit-interest-rates"),
    ("ICICI Bank", [
        "https://www.bankbazaar.com/fixed-deposit/icici-fixed-deposit-rate.html",
        "https://www.policybazaar.com/fd-interest-rates/icici-bank-fd-rates/",
        "https://www.paisabazaar.com/fixed-deposit/icici-bank-fd-rates/",
    ], "https://www.icicibank.com/personal-banking/deposits/fixed-deposit/fd-interest-rates"),
    ("Axis Bank", [
        "https://www.bankbazaar.com/fixed-deposit/axis-fixed-deposit-rate.html",
        "https://www.policybazaar.com/fd-interest-rates/axis-bank-fd-rates/",
        "https://www.paisabazaar.com/fixed-deposit/axis-bank-fd-rates/",
    ], "https://www.axisbank.com/fixed-deposit-interest-rate"),
    ("Kotak", [
        "https://www.bankbazaar.com/fixed-deposit/kotak-mahindra-bank-fixed-deposit-rate.html",
        "https://www.policybazaar.com/fd-interest-rates/kotak-mahindra-bank-fd-rates/",
        "https://www.paisabazaar.com/fixed-deposit/kotak-mahindra-bank-fd-rates/",
    ], "https://www.kotak.com/en/personal-banking/deposits/fixed-deposit/fixed-deposit-interest-rate.html"),
    ("AU SFB", [
        "https://www.policybazaar.com/fd-interest-rates/au-small-finance-bank-fd-rates/",
        "https://www.bankbazaar.com/fixed-deposit/au-small-finance-bank-fixed-deposit-rate.html",
        "https://www.paisabazaar.com/fixed-deposit/au-small-finance-bank-fd-rates/",
    ], "https://www.aubank.in/interest-rates/fixed-deposit-interest-rates"),
    ("Ujjivan SFB", [
        "https://www.policybazaar.com/fd-interest-rates/ujjivan-small-finance-bank-fd-rates/",
        "https://www.bankbazaar.com/fixed-deposit/ujjivan-small-finance-bank-fixed-deposit-rate.html",
        "https://www.paisabazaar.com/fixed-deposit/ujjivan-small-finance-bank-fd-rates/",
    ], "https://www.ujjivansfb.in/fixed-deposit-interest-rates"),
    ("Utkarsh SFB", [
        "https://www.policybazaar.com/fd-interest-rates/utkarsh-small-finance-bank-fd-rates/",
        "https://www.bankbazaar.com/fixed-deposit/utkarsh-small-finance-bank-fixed-deposit-rate.html",
        "https://www.paisabazaar.com/fixed-deposit/utkarsh-small-finance-bank-fd-rates/",
    ], "https://www.utkarsh.bank/fixed-deposit"),
    ("Unity SFB", [
        "https://www.policybazaar.com/fd-interest-rates/unity-small-finance-bank-fd-rates/",
        "https://www.bankbazaar.com/fixed-deposit/unity-small-finance-bank-fixed-deposit-rate.html",
    ], "https://theunitybank.com/fixed-deposit"),
    ("Suryoday SFB", [
        "https://www.policybazaar.com/fd-interest-rates/suryoday-small-finance-bank-fd-rates/",
        "https://www.bankbazaar.com/fixed-deposit/suryoday-small-finance-bank-fixed-deposit-rate.html",
    ], "https://www.suryodaybank.com/fixed-deposit"),
]


@st.cache_data(ttl=21600)
def fetch_all_fd():
    """Return list of (name, {tenor: rate}) tuples, fully automated.

    Stage 1 (parallel, HTTP): scrape every aggregator URL for each bank and
             MERGE the results, so a tenor missing on one site is filled by
             another. Thread-safe (no Streamlit cache calls, no Playwright).
    Stage 2 (sequential, MAIN thread): for any bank still missing tenors, open
             its aggregator page and then its own site in a headless browser to
             fill the gaps (handles JS-rendered tables that HTTP can't see)."""
    import concurrent.futures
    aggs_map = {n: aggs for n, aggs, _d in _FD_INSTITUTIONS}
    direct_map = {n: d for n, _a, d in _FD_INSTITUTIONS}
    by_name = {}

    # Stage 1 — all aggregator URLs, in parallel, merged per bank
    def _collect_http(item):
        name, aggs, _direct = item
        res = {}
        for u in aggs:
            _merge_tenors(res, _scrape_http(u))
            if len(res) == 5:
                break
        return name, res
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as ex:
        for name, res in ex.map(_collect_http, _FD_INSTITUTIONS):
            by_name[name] = res

    # Stage 2 — browser fill on the main thread, only for incomplete banks
    for name, aggs, direct in _FD_INSTITUTIONS:
        if len(by_name.get(name, {})) >= 5:
            continue
        for u in ([aggs[0]] if aggs else []) + ([direct] if direct else []):
            _merge_tenors(by_name[name], _scrape_browser(u))
            if len(by_name[name]) == 5:
                break

    return [(name, by_name[name]) for name, _a, _d in _FD_INSTITUTIONS]


@st.cache_data(ttl=3600)  # 1h — yields move daily; hourly is plenty
def bond_yield_browser(url):
    """Read a government-bond yield from an Investing.com page via headless
    browser. Strategy: the live yield is the prominent price near the top.
    We try a few known selectors for the 'last price' element, then fall back
    to the number that appears just before 'Prev. Close'. Returns float or None.
    """
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return None
    sels = ['[data-test="instrument-price-last"]',
            '[class*="instrument-price_last"]',
            '[class*="last_last"]']
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(user_agent=_BROWSER_H["User-Agent"])
            page.goto(url, timeout=45000, wait_until="domcontentloaded")
            page.wait_for_timeout(5000)
            val = None
            for sel in sels:                       # try the precise element
                try:
                    el = page.query_selector(sel)
                    if el:
                        t = el.inner_text().strip().replace(",", "")
                        m = re.search(r"\d{1,2}\.\d{1,3}", t)
                        if m:
                            val = float(m.group()); break
                except Exception:
                    pass
            if val is None:                        # fallback: value before "Prev. Close"
                body = re.sub(r"\s+", " ", page.inner_text("body"))
                m = re.search(r"(\d{1,2}\.\d{1,3})\s+[\d.+%\-() ]*Prev\.?\s*Close", body, re.I)
                if not m:                          # last resort: first plausible number
                    m = re.search(r"\b(\d{1,2}\.\d{2,3})\b", body)
                if m:
                    val = float(m.group(1))
            browser.close()
        # sanity: government yields sit roughly 0–15%
        return val if (val is not None and 0 < val < 15) else None
    except Exception:
        return None


@st.cache_data(ttl=3600)  # 1h — yields move daily; hourly is plenty
def bond_yield_change_browser(url):
    """Like bond_yield_browser, but also reads the day's move. Returns a tuple
    (value, change) where change = last - previous close, in yield points
    (so -0.02 means the yield fell 2 bps today). Either part may be None.
    Used for the India/UK/Japan 10Y daily-change column (US uses treasury.gov).
    Strategy: scrape the live 'last' price for the value, then read the
    'Prev. Close' figure and subtract - that label is stable and unambiguous,
    which avoids confusing the change element with the change-percent element.
    """
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return (None, None)
    sels = ['[data-test="instrument-price-last"]',
            '[class*="instrument-price_last"]',
            '[class*="last_last"]']
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page(user_agent=_BROWSER_H["User-Agent"])
            page.goto(url, timeout=45000, wait_until="domcontentloaded")
            page.wait_for_timeout(5000)
            val = None
            for sel in sels:                       # try the precise element
                try:
                    el = page.query_selector(sel)
                    if el:
                        m = re.search(r"\d{1,2}\.\d{1,3}",
                                      el.inner_text().strip().replace(",", ""))
                        if m:
                            val = float(m.group()); break
                except Exception:
                    pass
            body = re.sub(r"\s+", " ", page.inner_text("body"))
            if val is None:                        # fallback: value before "Prev. Close"
                m = re.search(r"(\d{1,2}\.\d{1,3})\s+[\d.+%\-() ]*Prev\.?\s*Close",
                              body, re.I)
                if m:
                    val = float(m.group(1))
            pc = re.search(r"Prev\.?\s*Close\s*(\d{1,2}\.\d{1,3})", body, re.I)
            prev_close = float(pc.group(1)) if pc else None
            browser.close()
        val = val if (val is not None and 0 < val < 15) else None
        change = None
        if val is not None and prev_close is not None:
            d = round(val - prev_close, 3)
            if abs(d) < 1.0:                       # ignore an implausible daily move
                change = d
        return (val, change)
    except Exception:
        return (None, None)


# ---------------------------------------------------------------------------
# FORMAT HELPERS
# ---------------------------------------------------------------------------
def chg(pct, bps=False, inverse=False):
    if pct is None:
        return '<span class="muted">-</span>'
    up = pct >= 0
    good = (not up) if inverse else up
    color = "var(--pos)" if good else "var(--neg)"
    arrow = "\u25B2" if up else "\u25BC"
    unit = " bps" if bps else "%"
    body = f"{pct:+.0f}" if bps else f"{abs(pct):.2f}"
    return f'<span style="color:{color}">{arrow} {body}{unit}</span>'


def num(x, d=2, prefix=""):
    if x is None:
        return '<span class="muted">-</span>'
    return prefix + f"{{:,.{d}f}}".format(x)


def gap(note="no free feed"):
    return f'<span class="muted">- <small>({note})</small></span>'


def hcells(d, keys=("m3", "m6", "y1", "y5")):
    """3M / 6M / 1Y / 5Y change cells for a Yahoo record (gap where unavailable)."""
    if not d:
        return [gap() for _ in keys]
    return [chg(d[k]) if d.get(k) is not None else gap() for k in keys]


def _pcfmt(price, pct, prefix, dec):
    """Stacked cell: price on top, change below. Muted 'NA' when data missing."""
    if price is None or pct is None:
        return '<span class="muted">NA</span>'
    return (f'{prefix}{price:,.{dec}f}'
            f'<br><span style="font-size:.85em">{chg(pct)}</span>')


def pc(price, pct, prefix="", dec=2):
    return _pcfmt(price, pct, prefix, dec)


def pcells(d, labels, prefix="", dec=2):
    """Stacked price+change cells for a Yahoo record across the given horizons."""
    if not d:
        return ['<span class="muted">NA</span>' for _ in labels]
    b = d.get("base", {})
    return [pc(b.get(l), d.get(l), prefix, dec) for l in labels]


def _inr(x):
    """Format a rupee amount: no decimals for large values, 2 for small."""
    if x is None:
        return None
    return f"\u20b9{x:,.0f}" if abs(x) >= 1000 else f"\u20b9{x:,.2f}"


def inr_pc(usd_now, usd_base, fx_now, fx_base, factor=1.0):
    """INR price-then + INR-terms % change (stacked). `factor` converts the
    USD native unit to the target Indian unit (e.g. oz->g). Uses live USD/INR
    history, so it stays automated. Muted 'NA' if any piece is missing."""
    if usd_base is None or fx_now is None or fx_base is None:
        return '<span class="muted">NA</span>'
    inr_then = usd_base * fx_base * factor
    inr_now = usd_now * fx_now * factor
    pct = (inr_now / inr_then - 1) * 100
    dec = 0 if abs(inr_then) >= 1000 else 2
    return _pcfmt(inr_then, pct, "\u20b9", dec)


# Fields that CAN be auto-pulled from FRED (free). suffix notes the cadence.
FRED_SERIES = {
    "india_10y":  ("INDIRLTLT01STM", "%", " (monthly)"),
    "japan_10y":  ("IRLTLT01JPM156N", "%", " (monthly)"),
    "uk_10y":     ("IRLTLT01GBM156N", "%", " (monthly)"),
    "germany_10y": ("IRLTLT01DEM156N", "%", " (monthly)"),
    "india_cpi":  ("INDCPIALLMINMEI", "", " (index, monthly)"),
    "india_call": ("IRSTCI01INM156N", "%", " (monthly)"),  # call money / interbank
}


def auto_val(field, note="no free feed"):
    """Manual typed value wins; else FRED auto; else a labelled gap."""
    v = str(manual.get(field, "")).strip()
    if v:
        return f'{v}'
    if field in FRED_SERIES and FRED_API_KEY:
        sid, unit, suffix = FRED_SERIES[field]
        val = fred(sid, FRED_API_KEY)
        if val is not None:
            return f'{val:.2f}{unit}'
    return gap(note)


_REPORT = []  # filled fresh each run; consumed by the download buttons


def _strip(c):
    """Plain text from an HTML cell, for the report files. Keeps <br> as a
    newline so price-over-change cells stay two lines in the export."""
    if c is None:
        return ""
    t = re.sub(r"<br\s*/?>", "\n", str(c))
    t = _html.unescape(_TAGS.sub("", t))
    return "\n".join(re.sub(r"\s+", " ", ln).strip() for ln in t.split("\n")).strip()


def report_section(title):
    """Start a report section AND render its heading on the page."""
    _REPORT.append({"title": title, "tables": []})
    st.subheader(title)


def render_table(headers, rows):
    th = "".join(f"<th>{h}</th>" for h in headers)
    trs = ""
    for r in rows:
        trs += "<tr>" + "".join(f"<td>{c}</td>" for c in r) + "</tr>"
    st.markdown(f'<div class="dash-wrap"><table class="dash">'
                f'<thead><tr>{th}</tr></thead>'
                f'<tbody>{trs}</tbody></table></div>', unsafe_allow_html=True)
    if _REPORT:  # record a clean copy for the downloadable report
        _REPORT[-1]["tables"].append(
            {"headers": list(headers),
             "rows": [[_strip(c) for c in r] for r in rows]})


def _pdf_safe(s):
    """reportlab core fonts are latin-1; swap symbols and drop the rest."""
    s = (str(s).replace("\u20b9", "Rs ").replace("\u25b2", "+")
         .replace("\u25bc", "-").replace("\u2013", "-").replace("\u00b7", "."))
    return s.encode("latin-1", "replace").decode("latin-1")


def _pdf_markup(s):
    """Like _pdf_safe but converts the dashboard's coloured HTML spans into
    reportlab <font color> markup, so up moves print green and down moves red."""
    s = str(s)
    s = s.replace("color:var(--pos)", "C_POS").replace("color:var(--neg)", "C_NEG")
    s = re.sub(r'<span[^>]*C_POS[^>]*>', '<font color="#0a8f3c">', s)   # green
    s = re.sub(r'<span[^>]*C_NEG[^>]*>', '<font color="#c0392b">', s)   # red
    s = re.sub(r'<span class="muted"[^>]*>', '<font color="#9aa0a6">', s)
    s = re.sub(r'<span[^>]*>', '<font>', s)        # any other span -> plain font
    s = s.replace("</span>", "</font>")
    s = s.replace("<small>", "").replace("</small>", "")
    s = s.replace("<br>", "<br/>")
    return _pdf_safe(s)


def chart_table():
    """Monthly 5-year history of the 10Y govt-bond yield for India / US / Japan,
    for embedding in the downloaded report. Returns (dates, {country: [values]})
    or (None, {}) if no FRED key / no data. Monthly resolution keeps the export
    light and tidy (the on-screen chart shows US daily)."""
    if not FRED_API_KEY:
        return None, {}
    series = {}
    for label, sid in (("India", "INDIRLTLT01STM"), ("US", "DGS10"),
                       ("Japan", "IRLTLT01JPM156N")):
        h = fred_history(sid, FRED_API_KEY, years=5)
        if h is not None and len(h):
            series[label] = h.resample("MS").last()
    if not series:
        return None, {}
    df = pd.DataFrame(series).sort_index().tail(61)
    dates = [d.strftime("%b %Y") for d in df.index]
    cols = {c: [None if pd.isna(x) else round(float(x), 2) for x in df[c]]
            for c in df.columns}
    return dates, cols


def chart_png(dates, cols):
    """Render the 5Y yield chart to a PNG (matplotlib). Returns a BytesIO or None
    if matplotlib isn't installed (the PDF then simply omits the chart)."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception:
        return None
    fig, ax = plt.subplots(figsize=(7.4, 3.2), dpi=150)
    x = list(range(len(dates)))
    for c in cols:
        ys = [v if v is not None else float("nan") for v in cols[c]]
        line, = ax.plot(x, ys, label=c, linewidth=1.5)
        # quarterly (every 3rd month) dot markers + value labels, like on-screen
        for i in range(0, len(x), 3):
            yi = ys[i] if i < len(ys) else float("nan")
            if yi == yi:   # skip NaN
                ax.plot(i, yi, "o", color=line.get_color(), markersize=2.8)
                ax.annotate(f"{yi:.2f}", (i, yi), textcoords="offset points",
                            xytext=(0, 4), ha="center", fontsize=4.2,
                            color=line.get_color())
    step = max(1, len(dates) // 8)
    ax.set_xticks(x[::step])
    ax.set_xticklabels(dates[::step], rotation=45, ha="right", fontsize=7)
    ax.set_ylabel("10Y yield (%)", fontsize=8)
    ax.set_title("10Y government bond yield — 5-year trend (monthly)", fontsize=9)
    ax.legend(fontsize=8); ax.grid(alpha=.3)
    fig.tight_layout()
    buf = io.BytesIO(); fig.savefig(buf, format="png"); plt.close(fig)
    buf.seek(0)
    return buf


def market_summary(eq, fx, cm, gold):
    """Build a short, data-driven 'what matters today' summary from the live
    numbers. Returns a list of plain-text bullet lines. Everything is derived
    from the same figures shown in the tables — nothing hand-written."""
    lines = []

    def emv(tk):
        d = eq.get(tk)
        if d and d.get("d1") is not None:
            return d["price"], d["d1"]
        return None, None

    # India headline (Nifty + Sensex)
    npx, nmv = emv("^NSEI")
    _spx, smv = emv("^BSESN")
    if nmv is not None:
        seg = f"Nifty 50 at {npx:,.0f} ({nmv:+.2f}%)"
        if smv is not None:
            seg += f", Sensex {smv:+.2f}%"
        lines.append("India equities: " + seg + ".")

    # Leader / laggard on the day (index name + country in brackets)
    moves = [(name, region, eq[tk]["d1"]) for region, name, tk in EQUITIES
             if eq.get(tk) and eq[tk].get("d1") is not None]
    if moves:
        best = max(moves, key=lambda x: x[2])
        worst = min(moves, key=lambda x: x[2])
        lines.append(f"{best[0]} ({best[1]}) led ({best[2]:+.2f}%); "
                     f"{worst[0]} ({worst[1]}) lagged ({worst[2]:+.2f}%).")

    # Rupee
    u = fx.get("INR=X")
    if u and u.get("d1") is not None:
        lines.append(f"Rupee: USD/INR {u['price']:.2f} ({u['d1']:+.2f}% today).")

    # Gold (live Indian rate; day-move proxied by COMEX trend)
    g = (gold or {}).get("24k")
    gmv = cm.get("GC=F", {}).get("d1") if cm.get("GC=F") else None
    if g:
        seg = f"Gold 24K Rs {g * 10:,.0f}/10g"
        if gmv is not None:
            seg += f" ({gmv:+.2f}%)"
        lines.append(seg + ".")

    # What matters today: the largest absolute index move
    if moves:
        big = max(moves, key=lambda x: abs(x[2]))
        verb = "rose" if big[2] >= 0 else "fell"
        tone = ("a quiet session with small moves across the board"
                if abs(big[2]) < 0.5 else
                f"{big[0]} ({big[1]}) {verb} {abs(big[2]):.2f}%, the day's biggest move")
        lines.append("What matters today: " + tone + ".")

    return lines


def build_excel(report, meta, summary=None):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Dashboard"
    bold = Font(bold=True)
    r = 1
    c = ws.cell(r, 1, "Daily Market Dashboard"); c.font = Font(bold=True, size=14); r += 1
    for m in meta:
        ws.cell(r, 1, m); r += 1
    r += 1
    if summary:
        c = ws.cell(r, 1, "Today's summary"); c.font = Font(bold=True, size=12); r += 1
        for s in summary:
            cell = ws.cell(r, 1, "• " + s)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
        r += 1
    for sec in report:
        c = ws.cell(r, 1, sec["title"]); c.font = Font(bold=True, size=12); r += 1
        for tbl in sec["tables"]:
            for ci, h in enumerate(tbl["headers"], 1):
                ws.cell(r, ci, h).font = bold
            r += 1
            for row in tbl["rows"]:
                for ci, val in enumerate(row, 1):
                    cell = ws.cell(r, ci, val)
                    if isinstance(val, str) and "\n" in val:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                r += 1
            r += 1
    for col in ws.columns:
        w = max((len(str(cell.value)) for cell in col if cell.value is not None),
                default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 60)

    # 5-year yield chart on its own sheet (native Excel line chart)
    try:
        dates, cols = chart_table()
        if dates and cols:
            from openpyxl.chart import LineChart, Reference
            cs = wb.create_sheet("5Y Yield Chart")
            cs.cell(1, 1, "Month")
            for j, c in enumerate(cols, start=2):
                cs.cell(1, j, c)
            for i, d in enumerate(dates, start=2):
                cs.cell(i, 1, d)
                for j, c in enumerate(cols, start=2):
                    cs.cell(i, j, cols[c][i - 2])
            ch = LineChart()
            ch.title = "10Y govt bond yield - 5-year trend (monthly)"
            ch.y_axis.title = "10Y yield (%)"; ch.x_axis.title = "Month"
            ch.height = 9; ch.width = 20
            data = Reference(cs, min_col=2, max_col=1 + len(cols),
                             min_row=1, max_row=1 + len(dates))
            cats = Reference(cs, min_col=1, min_row=2, max_row=1 + len(dates))
            ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
            cs.add_chart(ch, "G2")
    except Exception:
        pass

    bio = io.BytesIO(); wb.save(bio); return bio.getvalue()


def build_pdf(report, meta, summary=None):
    bio = io.BytesIO()
    doc = SimpleDocTemplate(bio, pagesize=A4, topMargin=12 * mm,
                            bottomMargin=12 * mm, leftMargin=10 * mm,
                            rightMargin=10 * mm)
    ss = getSampleStyleSheet()
    cell_st = ParagraphStyle("c", fontName="Helvetica", fontSize=6.4, leading=7.4)
    head_st = ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=6.4,
                             leading=7.4)
    el = [Paragraph("Daily Market Dashboard", ss["Title"])]
    for m in meta:
        el.append(Paragraph(_pdf_safe(m), ss["Normal"]))
    el.append(Spacer(1, 8))

    # Today's summary box (data-driven "what matters today")
    if summary:
        el.append(Paragraph("Today's summary", ss["Heading2"]))
        bullet_st = ParagraphStyle("b", fontName="Helvetica", fontSize=8.5,
                                   leading=12, leftIndent=8, spaceAfter=2)
        for s in summary:
            el.append(Paragraph("&bull;&nbsp; " + _pdf_safe(s), bullet_st))
        el.append(Spacer(1, 10))

    def para(x, hdr=False):
        # colour up/down moves; newline (price over change) -> <br/> so cells stack
        return Paragraph(_pdf_markup(x).replace("\n", "<br/>"),
                         head_st if hdr else cell_st)

    def add_chart():
        try:
            dates, cols = chart_table()
            if dates and cols:
                png = chart_png(dates, cols)
                if png:
                    el.append(Spacer(1, 4))
                    el.append(Image(png, width=doc.width, height=doc.width * 0.42))
                    el.append(Spacer(1, 6))
        except Exception:
            pass

    for sec in report:
        el.append(Paragraph(_pdf_safe(sec["title"]), ss["Heading2"]))
        for tbl in sec["tables"]:
            headers = tbl["headers"]
            ncol = max(1, len(headers))
            data = [[para(h, True) for h in headers]]
            data += [[para(x) for x in row] for row in tbl["rows"]]
            # fit the table to the printable width so nothing is cut off
            first = doc.width * (0.20 if ncol > 3 else 0.40)
            rest = (doc.width - first) / (ncol - 1) if ncol > 1 else doc.width
            widths = [first] + [rest] * (ncol - 1)
            t = Table(data, colWidths=widths, hAlign="LEFT", repeatRows=1)
            t.setStyle(TableStyle([
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#9db8ff")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1),
                 [colors.white, colors.HexColor("#f3f4f6")]),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d0d7de")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]))
            el.append(t); el.append(Spacer(1, 6))
        if sec["title"].lstrip().startswith("2"):   # chart sits inside section 2
            add_chart()
    doc.build(el); return bio.getvalue()


# ---------------------------------------------------------------------------
# PAGE
# ---------------------------------------------------------------------------
st.set_page_config(page_title="Daily Market Dashboard", layout="wide")
st.markdown("""<style>
:root{--pos:#26d07c;--neg:#ff5c6c;--accent:#5b8cff;--muted:#8b98a9}

/* page rhythm */
.block-container{padding-top:2.2rem;max-width:1180px}

/* title */
h1{font-weight:800!important;letter-spacing:-.02em;
  background:linear-gradient(92deg,#eaf1ff 0%,#9db8ff 60%,#5b8cff 100%);
  -webkit-background-clip:text;background-clip:text;-webkit-text-fill-color:transparent}

/* section headings get an accent bar */
h3{position:relative;padding-left:14px!important;margin-top:.4rem!important;
  font-weight:700!important;letter-spacing:-.01em}
h3::before{content:"";position:absolute;left:0;top:.18em;bottom:.18em;width:4px;
  border-radius:3px;background:linear-gradient(180deg,var(--accent),#9db8ff)}

/* table card */
.dash-wrap{border:1px solid rgba(255,255,255,.08);border-radius:14px;overflow:hidden;
  margin:6px 0 20px;background:linear-gradient(180deg,rgba(255,255,255,.035),rgba(255,255,255,.008));
  box-shadow:0 1px 2px rgba(0,0,0,.25)}
table.dash{width:100%;border-collapse:collapse;font-size:14px;margin:0}
table.dash thead th{text-align:right;color:var(--muted);font-weight:600;font-size:11px;
  letter-spacing:.05em;text-transform:uppercase;padding:11px 16px;
  background:rgba(255,255,255,.045);border-bottom:1px solid rgba(255,255,255,.10)}
table.dash td{text-align:right;padding:10px 16px;border-bottom:1px solid rgba(255,255,255,.05)}
table.dash tbody tr:nth-child(even){background:rgba(255,255,255,.022)}
table.dash tbody tr:hover{background:rgba(91,140,255,.10)}
table.dash tbody tr:last-child td{border-bottom:none}
table.dash th:first-child,table.dash td:first-child,
table.dash th:nth-child(2),table.dash td:nth-child(2){text-align:left}
table.dash td:first-child{font-weight:600}

/* management summary cards */
.mrow{display:flex;flex-wrap:wrap;gap:12px;margin:8px 0 6px}
.mcard{flex:1 1 150px;min-width:148px;padding:13px 16px;border-radius:13px;
  border:1px solid rgba(255,255,255,.08);
  background:linear-gradient(180deg,rgba(255,255,255,.045),rgba(255,255,255,.012))}
.mname{color:var(--muted);font-size:11px;letter-spacing:.06em;text-transform:uppercase;
  margin-bottom:5px}
.mval{font-size:19px;font-weight:700;line-height:1.2}
.msrc{color:#6b7280;font-size:10px;margin-top:5px}
.bw{display:flex;gap:18px;flex-wrap:wrap;margin:10px 0 2px;color:var(--muted);font-size:13px}

/* what-matters-today summary box */
.summary-box{margin:14px 0 4px;padding:14px 18px;border-radius:13px;
  border:1px solid rgba(91,140,255,.22);
  background:linear-gradient(180deg,rgba(91,140,255,.08),rgba(91,140,255,.02))}
.summary-h{font-size:11px;letter-spacing:.07em;text-transform:uppercase;
  color:var(--accent);font-weight:700;margin-bottom:7px}
.summary-list{margin:0;padding-left:18px}
.summary-list li{font-size:14px;line-height:1.6;margin:1px 0}

.muted{color:var(--muted)}
.src{color:var(--muted);font-size:12px;margin:-4px 0 12px}
</style>""", unsafe_allow_html=True)

with st.sidebar:
    st.header("Settings")
    secs = st.select_slider("Auto-refresh every (sec)",
                            [30, 60, 120, 300], value=60)
    st.caption("Leave this tab open; it refreshes itself.")
    if st.button("Refresh now"):
        st.cache_data.clear(); st.rerun()

    with st.expander("FRED key (one-time, enables US/EU/India macro)"):
        st.caption("Free key from fredaccount.stlouisfed.org/apikeys. Saved on "
                   "this computer. This is a login, not market data.")
        k = st.text_input("FRED API key", value=str(manual.get("fred_key", "")))
        if st.button("Save key"):
            save_manual({"fred_key": k})
            st.success("Saved. Refreshing..."); st.rerun()

    with st.expander("India macro (GDP / WPI — update on release)"):
        st.caption("No free live API exists for these. Enter the latest published "
                   "figure (with period), e.g. '7.80% (Q4 FY26)'. Leave blank to "
                   "use the built-in latest value.")
        g_in = st.text_input("India GDP growth",
                             value=str(manual.get("india_gdp", "")),
                             placeholder=f"{LATEST_INDIA_GDP[0]}% ({LATEST_INDIA_GDP[1]})")
        w_in = st.text_input("India WPI inflation (YoY)",
                             value=str(manual.get("india_wpi", "")),
                             placeholder=f"{LATEST_INDIA_WPI[0]}% ({LATEST_INDIA_WPI[1]})")
        if st.button("Save India macro"):
            save_manual({"india_gdp": g_in.strip(), "india_wpi": w_in.strip()})
            st.success("Saved. Refreshing..."); st.rerun()


if HAS_AUTOREFRESH:
    st_autorefresh(interval=secs*1000, key="auto")

st.title("Daily Market Dashboard")
st.caption(f"Last updated {dt.datetime.now():%a %d %b %Y  %H:%M:%S} {TZ}  ·  "
           "every number labelled with its source")

# fetch
_REPORT.clear()  # rebuild the report fresh every run
eq = yahoo(EQUITIES); fx = yahoo(FX); vl = yahoo(VOL)
cm = yahoo([("c", t) for t in dict.fromkeys(r[1] for r in COMMODITIES)])
alt = {**yahoo(REITS), **yahoo(INVITS)}
cry = yahoo(CRYPTO)
_gold = india_gold()      # live Indian retail 24K per gram (incl. duty/GST)
try:
    ty = treasury()
except Exception:
    ty = {"1Y": (None, None), "2Y": (None, None), "10Y": (None, None), "date": "-"}

# ---- Management summary ----
st.subheader("Management summary")
def mcard(name, value, change_html, src):
    return (f'<div class="mcard"><div class="mname">{name}</div>'
            f'<div class="mval">{value} {change_html}</div>'
            f'<div class="msrc">[{src}]</div></div>')
def hl(name, tk, src, d):
    v = d.get(tk)
    return mcard(name, num(v['price']) if v else '-',
                 chg(v['d1']) if v else '', src)
cards = []
if eq.get("^NSEI"): cards.append(hl("Nifty 50", "^NSEI", "Y", eq))
if eq.get("^GSPC"): cards.append(hl("S&P 500", "^GSPC", "Y", eq))
if ty["10Y"][0] is not None:
    cards.append(mcard("US 10Y", f"{ty['10Y'][0]:.2f}%",
                       chg(ty['10Y'][1]*100, bps=True, inverse=True), "T"))
if cm.get("GC=F"): cards.append(hl("Gold $/oz", "GC=F", "Y", cm))
if fx.get("INR=X"): cards.append(hl("USD/INR", "INR=X", "Y", fx))
if cards:
    st.markdown('<div class="mrow">' + "".join(cards) + '</div>',
                unsafe_allow_html=True)
perf = [(n, eq[t]["d1"]) for _, n, t in EQUITIES if eq.get(t)]
if perf:
    b = max(perf, key=lambda x: x[1]); w = min(perf, key=lambda x: x[1])
    st.markdown(f'<div class="bw"><span>Best: <b>{b[0]}</b> {chg(b[1])}</span>'
                f'<span>Worst: <b>{w[0]}</b> {chg(w[1])}</span></div>',
                unsafe_allow_html=True)

# data-driven "What matters today" summary (same lines used in the PDF/Excel)
try:
    _summary = market_summary(eq, fx, cm, _gold)
except Exception:
    _summary = []
if _summary:
    _items = "".join(f"<li>{_html.escape(s)}</li>" for s in _summary)
    st.markdown(
        '<div class="summary-box"><div class="summary-h">What matters today</div>'
        f'<ul class="summary-list">{_items}</ul></div>',
        unsafe_allow_html=True)
st.divider()

# ---- 1. Equities ----
report_section("1 · Global equity markets")
rows = []
for region, name, tk in EQUITIES:
    d = eq.get(tk)
    cur = num(d["price"]) if d else gap("fetch failed")
    rows.append([region, name, cur] + pcells(d, ("d1", "m1", "m3", "m6", "y1", "ytd")))
render_table(["Region", "Index", "Current", "1D", "1M", "3M", "6M", "1Y", "YTD"], rows)

# ---- 2. Rates ----
report_section("2 · Interest rates & fixed income")
YIELD_PAGES = {
    "india_1y":   "https://www.investing.com/rates-bonds/india-1-year-bond-yield",
    "india_2y":   "https://www.investing.com/rates-bonds/india-2-year-bond-yield",
    "uk_1y":      "https://www.investing.com/rates-bonds/uk-1-year-bond-yield",
    "uk_2y":      "https://www.investing.com/rates-bonds/uk-2-year-bond-yield",
    "japan_1y":   "https://www.investing.com/rates-bonds/japan-1-year-bond-yield",
    "japan_2y":   "https://www.investing.com/rates-bonds/japan-2-year-bond-yield",
}
_yld = {}
for k, u in YIELD_PAGES.items():
    _yld[k] = bond_yield_browser(u)

# 10Y pages give us a live value AND a daily change (last - prev close) for the
# three scraped countries, so the "Daily change (10Y)" column isn't US-only.
YIELD_10Y_PAGES = {
    "india": "https://www.investing.com/rates-bonds/india-10-year-bond-yield",
    "uk":    "https://www.investing.com/rates-bonds/uk-10-year-bond-yield",
    "japan": "https://www.investing.com/rates-bonds/japan-10-year-bond-yield",
}
_yld10 = {}
for k, u in YIELD_10Y_PAGES.items():
    _yld10[k] = bond_yield_change_browser(u)

def ycell(key):
    v = _yld.get(key)
    # Browser-scraped (Investing.com) — works locally, not on Streamlit Cloud
    # (no headless browser there). Show an honest gap rather than "scrape failed".
    return f"{v:.2f}%" if v is not None else gap("needs browser / no free feed")

def ten_y_cell(country):
    """(10Y value cell, daily-change cell). Uses the live scraped value+change;
    falls back to the FRED monthly value (change blank) if the scrape misses, so
    the row degrades gracefully instead of going empty."""
    v, c = _yld10.get(country, (None, None))
    val_cell = (f"{v:.2f}%" if v is not None
                else auto_val(f"{country}_10y"))
    chg_cell = chg(c * 100, bps=True, inverse=True) if c is not None else ""
    return val_cell, chg_cell

# Japan: prefer the MOF CSV (HTTP, cloud-friendly) over the browser scrape.
_jgb = japan_jgb()


def _jp_cell(key):
    v = _jgb.get(key)
    if v is not None:
        return f"{v:.2f}%"
    return ycell(f"japan_{key.lower()}")   # fall back to browser scrape if any


_in10, _inchg = ten_y_cell("india")
_uk10, _ukchg = ten_y_cell("uk")
# Japan 10Y + daily change from MOF if available, else browser, else FRED.
if _jgb.get("10Y") is not None:
    _jp10 = f"{_jgb['10Y']:.2f}%"
    _jpchg = (chg(_jgb["10Y_chg"] * 100, bps=True, inverse=True)
              if _jgb.get("10Y_chg") is not None else "")
else:
    _jp10, _jpchg = ten_y_cell("japan")

# Germany = euro-area AAA curve (ECB, daily, cloud-friendly), 10Y falls back to FRED.
_ecb = ecb_yield()


def _de_cell(key):
    v = _ecb.get(key)
    if v is not None:
        return f"{v:.2f}%"
    return auto_val("germany_10y") if key == "10Y" else gap("no free feed")


render_table(["Country", "1Y", "2Y", "10Y"], [
    ["India", ycell("india_1y"), ycell("india_2y"), _in10],
    ["US", num(ty["1Y"][0], 2) + "%" if ty["1Y"][0] else gap(),
     num(ty["2Y"][0], 2) + "%" if ty["2Y"][0] else gap(),
     num(ty["10Y"][0], 2) + "%" if ty["10Y"][0] else gap()],
    ["UK", ycell("uk_1y"), ycell("uk_2y"), _uk10],
    ["Germany", _de_cell("1Y"), _de_cell("2Y"), _de_cell("10Y")],
    ["Japan", _jp_cell("1Y"), _jp_cell("2Y"), _jp10],
])
st.markdown('<div class="src">US = Treasury.gov; Germany = ECB euro-area AAA curve; '
            'Japan = MOF JGB curve; India &amp; UK 10Y = FRED. India/UK 1Y/2Y come '
            'from a headless browser (work locally, blank on Streamlit Cloud — no '
            'free HTTP feed for those short-tenor yields).</div>',
            unsafe_allow_html=True)

# ---- 5-year trend: 10Y government bond yield, India / US / UK / Japan ----
# US comes back daily (DGS10); India/UK/Japan are FRED's monthly OECD series.
# Each line connects only its own real observations - nothing interpolated/faked.
TENY_SERIES = [
    ("US",      "DGS10",           "daily"),
    ("India",   "INDIRLTLT01STM",  "monthly"),
    ("UK",      "IRLTLT01GBM156N", "monthly"),
    ("Germany", "IRLTLT01DEM156N", "monthly"),
    ("Japan",   "IRLTLT01JPM156N", "monthly"),
]
_ORDER = ["India", "US", "UK", "Germany", "Japan"]
if FRED_API_KEY:
    _cut5 = pd.Timestamp(dt.date.today() - pd.Timedelta(days=365 * 5))
    _frames, _mark_frames = [], []
    for _label, _sid, _ in TENY_SERIES:
        _h = fred_history(_sid, FRED_API_KEY, years=5)
        if _h is not None and len(_h):
            _hh = _h[_h.index >= _cut5]
            _d = _hh.reset_index()
            _d.columns = ["date", "yield"]
            _d["country"] = _label
            _frames.append(_d)
            # one marker every 3 months (quarter starts), labelled with the value
            _q = _hh.resample("3MS").last().dropna()
            _m = _q.reset_index()
            _m.columns = ["date", "yield"]
            _m["country"] = _label
            _mark_frames.append(_m)
    if _frames:
        _long = pd.concat(_frames, ignore_index=True)
        _marks = pd.concat(_mark_frames, ignore_index=True)
        try:
            import altair as _altchart
            _x = _altchart.X("date:T", title=None,
                             axis=_altchart.Axis(format="%b %y", labelAngle=-45,
                                                 tickCount={"interval": "month",
                                                            "step": 3}))
            _color = _altchart.Color("country:N", title=None, sort=_ORDER)
            _tip = ["country:N", "date:T", _altchart.Tooltip("yield:Q", format=".2f")]
            _line = (_altchart.Chart(_long).mark_line(strokeWidth=1.6)
                     .encode(x=_x,
                             y=_altchart.Y("yield:Q", title="10Y yield (%)",
                                           scale=_altchart.Scale(zero=False)),
                             color=_color, tooltip=_tip))
            _dots = (_altchart.Chart(_marks)
                     .mark_point(filled=True, size=26, opacity=1)
                     .encode(x=_x, y="yield:Q", color=_color, tooltip=_tip))
            _labels = (_altchart.Chart(_marks)
                       .mark_text(dy=-9, fontSize=8, fontWeight="bold")
                       .encode(x=_x, y="yield:Q", color=_color,
                               text=_altchart.Text("yield:Q", format=".2f")))
            _chart = (_line + _dots + _labels).properties(height=420)
            st.altair_chart(_chart, use_container_width=True)
        except Exception:
            # Fallback if Altair is unavailable: forward-fill monthly across days
            _wide = (_long.pivot(index="date", columns="country", values="yield")
                          .sort_index().ffill())
            st.line_chart(_wide, height=420)
    else:
        st.markdown('<div class="src">5-year chart: no data returned from FRED '
                    '(try "Refresh now", or check the key).</div>',
                    unsafe_allow_html=True)
else:
    st.markdown('<div class="src">Add a free FRED key in the sidebar to see the '
                '5-year yield chart.</div>', unsafe_allow_html=True)

with st.expander("Yield scrape details (verify the picked value)"):
    st.caption("India/UK/Japan 1Y, 2Y & 10Y come from Investing.com via headless "
               "browser. A gap means the page changed or the browser is blocked.")
    for k, u in YIELD_PAGES.items():
        v = _yld.get(k)
        st.markdown(f"**{k}** — {v:.2f}% from {u}" if v is not None
                    else f"**{k}** — no value read from {u}")
    for k, u in YIELD_10Y_PAGES.items():
        v, c = _yld10.get(k, (None, None))
        if v is not None:
            cs = (f", daily change {c:+.2f} ({c*100:+.0f} bps)" if c is not None
                  else ", daily change n/a (no prev-close read)")
            st.markdown(f"**{k}_10y** — {v:.2f}%{cs} from {u}")
        else:
            st.markdown(f"**{k}_10y** — no value read (FRED monthly used) from {u}")
ff = fred("FEDFUNDS", FRED_API_KEY)
rbi = bis_policy("IN")
render_table(["Central bank", "Policy rate"], [
    ["RBI Repo Rate", rbi if rbi else gap("BIS feed unavailable")],
    ["US Fed Funds Rate", f"{ff:.2f}%" if ff is not None else gap("add FRED key")],
])

# ---- 3. Bank & SFB deposit rates ----
report_section("3 · Bank & SFB deposit rates (India)")
# Rates are scraped live from PolicyBazaar (raw HTML) every 6 hours.
# If a cell shows NA the page layout may have changed or was blocked —
# the cache will be retried on the next auto-refresh.

def _fd(v):
    return f"{v}%" if str(v).strip() else '<span class="muted">NA</span>'


with st.spinner("Loading FD rates…"):
    _fd_data = fetch_all_fd()

fd_rows = []
for name, rates in _fd_data:
    fd_rows.append([
        name,
        _fd(rates.get("1M", "")),
        _fd(rates.get("3M", "")),
        _fd(rates.get("6M", "")),
        _fd(rates.get("1Y", "")),
        _fd(rates.get("3Y", "")),
    ])

render_table(["Institution", "1M", "3M", "6M", "1Y", "3Y"], fd_rows)
st.markdown('<div class="src">Source: BankBazaar / PolicyBazaar / Paisabazaar + '
            'bank sites (scraped &amp; merged across sources, cached 6h). '
            'Verify at the issuer before investing.</div>',
            unsafe_allow_html=True)

with st.expander("FD scrape details (which banks filled / are missing tenors)"):
    st.caption("Each bank is scraped from multiple aggregators and its own site, "
               "then merged. A bank showing fewer than 5 tenors means those "
               "pages changed, blocked the request, or render rates via JS the "
               "browser couldn't reach. Hit “Refresh now” to retry.")
    for name, rates in _fd_data:
        got = [t for t in ("1M", "3M", "6M", "1Y", "3Y") if rates.get(t)]
        miss = [t for t in ("1M", "3M", "6M", "1Y", "3Y") if not rates.get(t)]
        if len(got) == 5:
            st.markdown(f"✅ **{name}** — all 5 tenors")
        elif got:
            st.markdown(f"⚠️ **{name}** — got {', '.join(got)}; "
                        f"missing {', '.join(miss)}")
        else:
            st.markdown(f"❌ **{name}** — nothing read")

# ---- 4. Currencies ----
report_section("4 · Currency markets")
rows = []
for name, tk in FX:
    d = fx.get(tk)
    cur = num(d["price"], 4) if d else gap()
    rows.append([name, cur] + pcells(d, ("d1", "m1", "m3", "m6", "y1", "y5"), "", 4))
render_table(["Pair", "Current", "1D", "1M", "3M", "6M", "1Y", "5Y"], rows)

# ---- 5. Commodities ----
report_section("5 · Commodities (INR)")
_inrfx = fx.get("INR=X")
_fxnow = _inrfx["price"] if _inrfx else None
_fxbase = _inrfx.get("base", {}) if _inrfx else {}
# _gold fetched once in the main fetch block above (live Indian retail 24K/g)


def _gold_cell(indian_now, pct):
    """A 3M/6M/etc. cell for gold on the INDIAN-retail scale. We anchor to the
    live Indian price and walk back by the international % move for that horizon,
    so the price-then shown is on the same scale as 'Current' (no COMEX/Indian
    scale mismatch). The % itself reflects the international gold trend."""
    if indian_now is None or pct is None:
        return '<span class="muted">NA</span>'
    then = indian_now / (1 + pct / 100.0)
    dec = 0 if abs(then) >= 1000 else 2
    return _pcfmt(then, pct, "\u20b9", dec)


rows = []
for name, tk, factor in COMMODITIES:
    d = cm.get(tk)
    if not d or _fxnow is None:
        rows.append([name, gap("fetch failed")] + ['<span class="muted">NA</span>'] * 5)
        continue

    # Gold 24K: use the LIVE Indian retail rate for Current, and make the change
    # columns consistent with it (anchored to the Indian price, % from COMEX).
    if name.startswith("Gold 24K"):
        mult = 10 if "10g" in name else 1
        g = _gold.get("24k")
        if g:
            indian_now = g * mult
            cur = _inr(indian_now)
            cells = [_gold_cell(indian_now, d.get(lbl))
                     for lbl in ("d1", "m3", "m6", "y1", "y5")]
        else:
            # scrape failed → fall back to COMEX-spot conversion so it's not blank
            cur = _inr(d["price"] * _fxnow * factor)
            b = d.get("base", {})
            cells = [inr_pc(d["price"], b.get(lbl), _fxnow, _fxbase.get(lbl), factor)
                     for lbl in ("d1", "m3", "m6", "y1", "y5")]
        rows.append([name, cur] + cells)
        continue

    # Everything else: international spot converted to INR (unchanged).
    cur = _inr(d["price"] * _fxnow * factor)
    b = d.get("base", {})
    cells = [inr_pc(d["price"], b.get(lbl), _fxnow, _fxbase.get(lbl), factor)
             for lbl in ("d1", "m3", "m6", "y1", "y5")]
    rows.append([name, cur] + cells)
render_table(["Commodity", "Current", "1D", "3M", "6M", "1Y", "5Y"], rows)
st.markdown('<div class="src">Gold 24K = live Indian retail rate (Goodreturns, '
            'incl. duty &amp; GST); change columns track the international gold '
            'trend. Silver/crude/metals = international spot converted to INR.</div>',
            unsafe_allow_html=True)

# ---- 6. Alt assets ----
report_section("6 · Alternative assets")
rows = []
for name, tk in CRYPTO:
    d = cry.get(tk)
    cur = num(d["price"], 0, "$") if d else gap()
    rows.append([name, cur] + pcells(d, ("d1", "m1", "m3", "m6", "y1", "y5"), "$", 0))
for label, tk in REITS + INVITS:
    d = alt.get(tk)
    cur = num(d["price"], 2, "\u20b9") if (d and d.get("price")) else gap("fetch failed")
    rows.append([label, cur] +
                pcells(d if (d and d.get("price")) else None,
                       ("d1", "m1", "m3", "m6", "y1", "y5"), "\u20b9", 2))
render_table(["Asset", "Current", "1D", "1M", "3M", "6M", "1Y", "5Y"], rows)

# ---- 7. Volatility ----
report_section("7 · Volatility & risk indicators")
ivd = vl.get("^INDIAVIX")
iv = (ivd["price"], ivd["d1"]) if ivd else None
render_table(["Indicator", "Current", "Change"], [
    ["India VIX", num(iv[0]) if iv else gap(), chg(iv[1], inverse=True) if iv else ""],
])

# ---- 8. Macro ----
report_section("8 · Macro indicators")
cpi = fred("CPIAUCSL", FRED_API_KEY); un = fred("UNRATE", FRED_API_KEY)
# Forex reserves: prefer FRED monthly (current, excl. gold); fall back to the
# World Bank annual (incl. gold, but lags by ~1-2 years) if no FRED key.
_resv, _resv_d = fred_latest("TRESEGINM052N", FRED_API_KEY)   # USD millions, monthly
if _resv is not None:
    _resv_cell = f"${_resv/1000:,.2f}B ({_resv_d}, excl. gold)"
else:
    fxr = worldbank("FI.RES.TOTL.CD")
    _resv_cell = (f"${fxr[0]/1e9:,.2f}B ({fxr[1]}, incl. gold)" if fxr
                  else gap("no free source"))
# India GDP growth & WPI inflation have no free real-time API — show the latest
# published figure (manual override in the sidebar wins, else the seeded default).
_gdp_m = str(manual.get("india_gdp", "")).strip()
_wpi_m = str(manual.get("india_wpi", "")).strip()
_gdp_cell = _gdp_m if _gdp_m else f"{LATEST_INDIA_GDP[0]}% ({LATEST_INDIA_GDP[1]})"
_wpi_cell = _wpi_m if _wpi_m else f"{LATEST_INDIA_WPI[0]}% ({LATEST_INDIA_WPI[1]})"
render_table(["Indicator", "Latest"], [
    ["India CPI", auto_val("india_cpi")],
    ["India WPI inflation (YoY)", _wpi_cell],
    ["India GDP Growth", _gdp_cell],
    ["US CPI Index", f"{cpi:.2f}" if cpi is not None else gap("add FRED key")],
    ["US Unemployment %", f"{un:.2f}" if un is not None else gap("add FRED key")],
    ["India Forex Reserves", _resv_cell],
])
st.markdown('<div class="src">India GDP &amp; WPI are released only in periodic '
            'government press releases (no free live API); the latest published '
            'figure is shown and can be updated in the sidebar.</div>',
            unsafe_allow_html=True)

# ---- footer ----
st.divider()
st.caption("Verify figures at the source.")

# ---- downloadable daily report (built from everything rendered above) ----
_stamp = dt.datetime.now()
_meta = [f"Generated {_stamp:%a %d %b %Y  %H:%M:%S} {TZ}"]
_fname = f"market-dashboard-{_stamp:%Y-%m-%d}"
try:
    _summary = market_summary(eq, fx, cm, _gold)
except Exception:
    _summary = []
with st.sidebar:
    st.divider()
    st.markdown("**Download today's report**")
    if HAS_XLSX:
        try:
            st.download_button(
                "Excel (.xlsx)", data=build_excel(_REPORT, _meta, _summary),
                file_name=_fname + ".xlsx",
                mime="application/vnd.openxmlformats-officedocument."
                     "spreadsheetml.sheet")
        except Exception as e:
            st.caption(f"Excel error: {e}")
    else:
        st.caption("Excel needs the openpyxl library "
                   "(run: pip install openpyxl).")
    if HAS_PDF:
        try:
            st.download_button(
                "PDF (.pdf)", data=build_pdf(_REPORT, _meta, _summary),
                file_name=_fname + ".pdf", mime="application/pdf")
        except Exception as e:
            st.caption(f"PDF error: {e}")
    else:
        st.caption("PDF needs the reportlab library "
                   "(run: pip install reportlab).")
    st.caption("Click a button to download a dated snapshot of today's numbers.")
